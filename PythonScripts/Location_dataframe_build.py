#Location dataframe build - GoogleMapsAPI

import requests
from openpyxl import load_workbook
from dotenv import load_dotenv
import os
import time

#load environment variables to retrieve API key
load_dotenv()

#Your Google API Key
api_key = os.getenv('Google_Maps_API_Key')

wb = load_workbook(r"C:\Users\BSA-OliverJ'22\Projects\Database Architecture\ExcelWorkbooks\SQL_master_import_book.xlsx")
ws = wb['dbo.Location'] 

# Define column index positions
country_name_col_index = 7
iso2_col_index = 8
iso3_col_index = 9
latitude_col_index = 10
longitude_col_index = 11
state_name_col_index = 5
state_name_code_col_index = 6
city_name_col_index = 4
time_zone_col_index = 12

# need to import the list of locations here

for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
    hfm_location_name = row[0].value

    #Geocoding API request
    geocode_url = f"https://maps.googleapis.com/maps/api/geocode/json?address={hfm_location_name}&key={api_key}"
    response = requests.get(geocode_url)
    geocode_data = response.json()

    if geocode_data['status'] == 'OK':
        # Extracting data from the API response
        location_data = geocode_data['results'][0]
        #Extracting the country name
        country_name = None
        iso_code2 = None
        latitude = None
        longitude = None
        state_name = None
        state_code = None
        city_name = None

        # Extract CityName
        for component in location_data['address_components']:
            if 'locality' in component['types']:
                city_name = component['long_name']
                break
            elif 'postal_town' in component['types']:
                city_name = component['long_name']
                break
            elif 'administrative_area_level_2' in component['types']:
                city_name = component['long_name']
                break

        # Extract StateName and StateNameCode (if the location is in the United States)
        for component in location_data['address_components']:
            if 'country' in component['types'] and component['short_name'] == 'US':
                # The location is in the United States, extract state information
                for component in location_data['address_components']:
                    if 'administrative_area_level_1' in component['types']:
                        state_name = component['long_name']
                        state_code = component['short_name']
                        break
    
        for component in location_data['address_components']:
            if 'country' in component['types']:
                country_name = component['long_name']
                iso_code2 = component['short_name']

        
        # Extract latitude and longitude
        latitude = location_data['geometry']['location']['lat']
        longitude = location_data['geometry']['location']['lng']

        # Time Zone API request
        timestamp = int(time.time())  # current time
        timezone_url = f"https://maps.googleapis.com/maps/api/timezone/json?location={latitude},{longitude}&timestamp={timestamp}&key={api_key}"
        timezone_response = requests.get(timezone_url)
        timezone_data = timezone_response.json()

        if timezone_data['status'] == 'OK':
            # Extract the timezone information
            timezone_id = timezone_data['timeZoneId']  # e.g., 'America/Los_Angeles'

            # Write the TimeZoneID into the Excel workbook
            ws.cell(row=row[0].row, column=time_zone_col_index).value = timezone_id
        else:
            print(f"Time Zone API request failed for {hfm_location_name} with status: {timezone_data['status']}.")
    
        # Print the data to the terminal (optional)
        print(f"Data for {hfm_location_name}: Country: {country_name}, ISO2: {iso_code2}, Latitude: {latitude}, Longitude: {longitude}")

        # Write the data into the Excel workbook
        ws.cell(row=row[0].row, column=country_name_col_index).value = country_name
        ws.cell(row=row[0].row, column=iso2_col_index).value = iso_code2
        ws.cell(row=row[0].row, column=latitude_col_index).value = latitude
        ws.cell(row=row[0].row, column=longitude_col_index).value = longitude
        ws.cell(row=row[0].row, column=state_name_col_index).value = state_name
        ws.cell(row=row[0].row, column=state_name_code_col_index).value = state_code
        ws.cell(row=row[0].row, column=city_name_col_index).value = city_name
        # ISO3 code logic will be added here once you have the mapping or a service to get it
        
    else:
        print(f"Geocoding API request failed for {hfm_location_name} with status: {geocode_data['status']}.")

# Save the workbook
wb.save(r"C:\Users\BSA-OliverJ'22\Projects\Database Architecture\ExcelWorkbooks\SQL_master_import_book.xlsx")

# Close the workbook if it's no longer needed
wb.close()



